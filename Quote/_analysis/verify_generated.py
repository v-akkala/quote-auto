from pathlib import Path

from openpyxl import load_workbook

path = Path(r"C:\Users\Vikram\Desktop\Coding Stuff\SKC\Quote\outputs\SQ26270001.xlsx")
wb = load_workbook(path, data_only=False)
ws = wb.active
print(f"sheet={ws.title}")
for cell in ("AJ3", "AR4", "L6", "AJ6", "X10", "B14", "D14", "AE14", "AI14", "AO14", "AT14", "B15", "D15", "AE15", "AI15", "AO15", "AT15", "AN16", "AN20"):
    print(f"{cell}={ws[cell].value}")
print("merged_test_rows", [str(rng) for rng in ws.merged_cells.ranges if 13 <= rng.min_row <= 20 and rng.min_col <= 50])
