from pathlib import Path

from openpyxl import load_workbook

BOOKS = [
    Path(r"C:\Users\Vikram\Downloads\Template Quote (2).xlsx"),
    Path(r"C:\Users\Vikram\Downloads\Quote Copy 203.xlsx"),
]


def clean(value):
    return str(value).replace("\n", " | ")[:180]


for path in BOOKS:
    wb = load_workbook(path)
    ws = wb.active
    print(f"=== {path.name} :: {ws.title} :: {ws.max_row}x{ws.max_column}")
    merged = [
        str(cell_range)
        for cell_range in ws.merged_cells.ranges
        if cell_range.min_row <= 35 and cell_range.max_row >= 1
    ]
    print("MERGED", merged[:120])
    for row in range(1, 71):
        values = []
        for col in range(1, 61):
            cell = ws.cell(row, col)
            if cell.value not in (None, ""):
                values.append(f"{cell.coordinate}={clean(cell.value)}")
        if values:
            print(f"ROW {row}: " + "; ".join(values))
    print()
