from __future__ import annotations

import json
import re
import shutil
from copy import copy
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.cell_range import CellRange

ROOT_DIR = Path(__file__).resolve().parent.parent
TEMPLATE_PATH = ROOT_DIR / "templates" / "Template Quote.xlsx"
DISCOUNT_TEMPLATE_PATH = ROOT_DIR / "templates" / "Template Quote with Discount.xlsx"
SIGNATURE_STAMP_PATH = ROOT_DIR / "assets" / "signature_stamp.png"
OUTPUT_DIR = ROOT_DIR / "outputs"
DATA_DIR = ROOT_DIR / "data"
COUNTER_PATH = DATA_DIR / "quote_counter.json"

QUOTE_PREFIX = "SKC2627"
STARTING_QUOTE_NUMBER = 1
SAC_CODE = 998346
TEST_START_ROW = 14
TOTAL_START_ROW = 15
TEST_MERGES = ((2, 3), (4, 30), (31, 33), (35, 40), (41, 45), (46, 50))
FILENAME_SAFE_RE = re.compile(r"^[A-Za-z0-9_-]+$")


@dataclass
class TestItem:
    name: str
    requirements: str
    qty: str = "1"
    batch: str = "1"
    total_cost: str = ""
    standard: str = ""
    product_keys: list[str] = field(default_factory=list)


@dataclass
class ProductItem:
    name: str
    dim1: str = ""
    dim2: str = ""
    dim3: str = ""
    dim_unit: str = "mm"
    weight: str = ""
    weight_unit: str = "kg"


def today_string() -> str:
    now = date.today()
    return f"{now.month}-{now.day}-{now.year}"


def next_quote_ref() -> str:
    return f"{QUOTE_PREFIX}{next_quote_suffix()}"


def next_quote_suffix() -> str:
    return f"{_read_counter():04d}"


def _read_counter() -> int:
    if not COUNTER_PATH.exists():
        return STARTING_QUOTE_NUMBER
    try:
        data = json.loads(COUNTER_PATH.read_text(encoding="utf-8"))
        value = int(data.get("next_number", STARTING_QUOTE_NUMBER))
        return max(value, STARTING_QUOTE_NUMBER)
    except (ValueError, TypeError, json.JSONDecodeError):
        return STARTING_QUOTE_NUMBER


def _write_counter(next_number: int) -> None:
    DATA_DIR.mkdir(exist_ok=True)
    COUNTER_PATH.write_text(
        json.dumps({"next_number": next_number}, indent=2),
        encoding="utf-8",
    )


def reserve_quote_ref(suffix: str = "") -> str:
    clean_suffix = normalize_quote_suffix(suffix)
    if clean_suffix:
        _advance_counter_for_suffix(clean_suffix)
        return f"{QUOTE_PREFIX}{clean_suffix}"

    number = _read_counter()
    _write_counter(number + 1)
    return f"{QUOTE_PREFIX}{number:04d}"


def normalize_quote_suffix(suffix: str) -> str:
    clean_suffix = str(suffix or "").strip()
    if not clean_suffix:
        return ""
    if clean_suffix.upper().startswith(QUOTE_PREFIX):
        clean_suffix = clean_suffix[len(QUOTE_PREFIX) :]
    if not FILENAME_SAFE_RE.fullmatch(clean_suffix):
        raise ValueError("Quote number suffix can only contain letters, numbers, underscores, and hyphens.")
    return clean_suffix


def _advance_counter_for_suffix(suffix: str) -> None:
    if not suffix.isdigit():
        return
    used_number = int(suffix)
    current = _read_counter()
    if used_number >= current:
        _write_counter(used_number + 1)


def parse_date_text(value: str) -> str:
    value = value.strip()
    if not value:
        return today_string()

    for fmt in ("%m-%d-%Y", "%m/%d/%Y", "%d-%m-%Y", "%d/%m/%Y", "%Y-%m-%d"):
        try:
            parsed = datetime.strptime(value, fmt).date()
            return f"{parsed.month}-{parsed.day}-{parsed.year}"
        except ValueError:
            pass
    return value


def coerce_number(value: str) -> Any:
    text = str(value).strip()
    if not text:
        return ""
    if text.startswith("="):
        return text
    try:
        number = float(text.replace(",", ""))
    except ValueError:
        return text
    if number.is_integer():
        return int(number)
    return number


def parse_percent(value: str) -> float:
    text = str(value or "").strip().replace("%", "")
    if not text:
        return 0.0
    try:
        percent = float(text)
    except ValueError as exc:
        raise ValueError("Discount must be a number, such as 6 or 6%.") from exc
    if percent < 0 or percent > 100:
        raise ValueError("Discount must be between 0 and 100.")
    return percent


def is_truthy(value: Any) -> bool:
    return str(value).strip().lower() in {"1", "true", "yes", "y", "on"}


def parse_decimal_field(value: str, field_name: str, *, allow_blank: bool = False) -> Decimal | None:
    text = str(value or "").strip().replace(",", "")
    if not text:
        if allow_blank:
            return None
        raise ValueError(f"{field_name} is required.")
    if text.startswith("="):
        raise ValueError(f"{field_name} must be a number, not a formula.")
    try:
        number = Decimal(text)
    except InvalidOperation as exc:
        raise ValueError(f"{field_name} must be a number. Do not include currency text like Rs.") from exc
    if number < 0:
        raise ValueError(f"{field_name} cannot be negative.")
    return number


def _validate_test_numbers(tests: list[TestItem]) -> None:
    for index, test in enumerate(tests, start=1):
        label = test.name or test.standard or f"test {index}"
        parse_decimal_field(test.qty or "1", f"Sample Qty for {label}")
        parse_decimal_field(test.batch or "1", f"Test Qty for {label}")
        parse_decimal_field(test.total_cost, f"Total Cost for {label}", allow_blank=True)


def _validate_unique_product_names(products: list[ProductItem]) -> None:
    names = [product.name.strip() for product in products if product.name.strip()]
    duplicates = sorted({name for name in names if names.count(name) > 1})
    if duplicates:
        raise ValueError("Product names must be unique. Duplicate: " + ", ".join(duplicates))


def format_product_details(data: dict[str, str], products: list[ProductItem] | None = None) -> str:
    product_items = products or [
        ProductItem(
            name=data.get("product_name", "").strip(),
            dim1=data.get("dim1", "").strip(),
            dim2=data.get("dim2", "").strip(),
            dim3=data.get("dim3", "").strip(),
            dim_unit=data.get("dim_unit", "mm").strip() or "mm",
            weight=data.get("weight", "").strip(),
            weight_unit=data.get("weight_unit", "kg").strip() or "kg",
        )
    ]

    blocks = [_format_one_product(product) for product in product_items if product.name.strip()]
    if not blocks:
        return "Product Name: \nDimensions: \nWeight: "
    return "\n\n".join(blocks)


def _format_one_product(product: ProductItem) -> str:
    dimensions = "x".join(part.strip() for part in (product.dim1, product.dim2, product.dim3) if part.strip())
    if dimensions:
        dimensions = f"{dimensions} {(product.dim_unit or 'mm').strip()}"

    weight = product.weight.strip()
    weight_text = f"{weight} {(product.weight_unit or 'kg').strip()}" if weight else ""
    return (
        f"Product Name: {product.name.strip()}\n"
        f"Dimensions: {dimensions}\n"
        f"Weight: {weight_text}"
    )


def product_key(product: ProductItem) -> str:
    return product.name.strip()


def product_quote_label(product: ProductItem) -> str:
    name = product.name.strip()
    dimensions = " x ".join(part.strip() for part in (product.dim1, product.dim2, product.dim3) if part.strip())
    if dimensions:
        dimensions = f"{dimensions}{(product.dim_unit or 'mm').strip()}"
    weight = product.weight.strip()
    weight_text = f"{weight}{(product.weight_unit or 'kg').strip()}" if weight else ""
    parts = [part for part in (name, dimensions, weight_text) if part]
    return " - ".join(parts)


def format_customer_details(data: dict[str, str]) -> str:
    return "\n".join(
        [
            f"Customer: {data.get('customer', '').strip()}",
            f"Designation: {data.get('designation', '').strip()}",
            f"Company: {data.get('company', '').strip()}",
            f"Tel No.: {data.get('tel_no', '').strip()}",
            f"E-mail: {data.get('email', '').strip()}",
            f"Address: {data.get('address', '').strip()}",
            "",
        ]
    )


def format_agency_details(existing: str, quote_made_by: str, technical_engineers: str) -> str:
    lines = str(existing or "").splitlines()
    found_quote_by = False
    found_engineers = False
    output = []
    for line in lines:
        stripped = line.strip()
        if stripped.startswith("Quote Made By:"):
            output.append(f"Quote Made By: {quote_made_by.strip()}")
            found_quote_by = True
        elif stripped.startswith("Technical Engineer(s):"):
            output.append(f"Technical Engineer(s): {technical_engineers.strip()}")
            found_engineers = True
        else:
            output.append(line)
    if not found_quote_by:
        output.append(f"Quote Made By: {quote_made_by.strip()}")
    if not found_engineers:
        output.append(f"Technical Engineer(s): {technical_engineers.strip()}")
    return "\n".join(output)


def generate_quote(data: dict[str, str], tests: list[TestItem], products: list[ProductItem] | None = None) -> Path:
    use_discount = is_truthy(data.get("discount_enabled", ""))
    discount_percent = parse_percent(data.get("discount_percent", "")) if use_discount else 0.0
    template_path = DISCOUNT_TEMPLATE_PATH if use_discount else TEMPLATE_PATH
    if not template_path.exists():
        raise FileNotFoundError(f"Template not found: {template_path}")
    if not tests:
        raise ValueError("Add at least one test before generating a quote.")
    _validate_test_numbers(tests)
    if any(test.product_keys for test in tests):
        _validate_unique_product_names(products or [])
    product_lookup = {product_key(product): product for product in (products or [])}
    display_rows = build_test_display_rows(tests, product_lookup)

    OUTPUT_DIR.mkdir(exist_ok=True)
    quote_ref = reserve_quote_ref(data.get("quote_suffix", ""))
    output_path = OUTPUT_DIR / f"{quote_ref}.xlsx"
    shutil.copy2(template_path, output_path)

    workbook = load_workbook(output_path)
    worksheet = workbook.active
    worksheet.title = quote_ref

    worksheet["AJ3"] = quote_ref
    worksheet["AJ4"] = "Dated"
    worksheet["AR4"] = parse_date_text(data.get("quote_date", today_string()))
    worksheet["B6"] = format_agency_details(
        worksheet["B6"].value,
        data.get("quote_made_by", ""),
        data.get("technical_engineers", ""),
    )
    product_details = format_product_details(data, products)
    worksheet["L6"] = product_details
    _fit_product_rows(worksheet, product_details)
    worksheet["AJ6"] = format_customer_details(data)
    worksheet["X10"] = parse_date_text(data.get("email_date", today_string()))
    worksheet["AI13"] = "Sample Qty"
    worksheet["AO13"] = "Test Qty"

    _prepare_test_rows(worksheet, len(display_rows))
    _write_tests(worksheet, display_rows)
    _write_totals(
        worksheet,
        display_rows,
        use_discount=use_discount,
        discount_percent=discount_percent,
        instate=is_truthy(data.get("instate", "")),
    )
    _write_payment_terms(worksheet, data.get("payment_terms", ""))
    _ensure_signature_stamp(worksheet)

    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.save(output_path)
    return output_path


def _fit_product_rows(worksheet, product_details: str) -> None:
    line_count = max(3, product_details.count("\n") + 1)
    total_height = min(409, max(78, line_count * 15))
    worksheet.row_dimensions[6].height = total_height / 2
    worksheet.row_dimensions[7].height = total_height / 2


def _prepare_test_rows(worksheet, row_count: int) -> None:
    rows_to_insert = max(row_count - 1, 0)
    if rows_to_insert:
        shifted_merges = _remove_merges_for_insert(worksheet, TOTAL_START_ROW, rows_to_insert)
        worksheet.insert_rows(TOTAL_START_ROW, rows_to_insert)
        _shift_images_for_insert(worksheet, TOTAL_START_ROW, rows_to_insert)
        for merge_range in shifted_merges:
            worksheet.merge_cells(str(merge_range))

    for offset in range(row_count):
        row = TEST_START_ROW + offset
        _copy_row_style(worksheet, TEST_START_ROW, row)
        _ensure_test_row_merges(worksheet, row)
        worksheet.row_dimensions[row].height = max(worksheet.row_dimensions[row].height or 48, 48)


def _remove_merges_for_insert(worksheet, insert_at: int, amount: int) -> list[CellRange]:
    shifted = []
    for merge_range in list(worksheet.merged_cells.ranges):
        cell_range = CellRange(str(merge_range))
        if cell_range.min_row >= insert_at:
            worksheet.unmerge_cells(str(cell_range))
            cell_range.shift(row_shift=amount, col_shift=0)
            shifted.append(cell_range)
    return shifted


def _copy_row_style(worksheet, source_row: int, target_row: int) -> None:
    if source_row == target_row:
        return
    worksheet.row_dimensions[target_row].height = worksheet.row_dimensions[source_row].height
    for col in range(1, worksheet.max_column + 1):
        source = worksheet.cell(source_row, col)
        target = worksheet.cell(target_row, col)
        if source.has_style:
            target._style = copy(source._style)
        if source.number_format:
            target.number_format = source.number_format
        if source.alignment:
            target.alignment = copy(source.alignment)
        if source.font:
            target.font = copy(source.font)
        if source.fill:
            target.fill = copy(source.fill)
        if source.border:
            target.border = copy(source.border)
        if source.protection:
            target.protection = copy(source.protection)


def _ensure_test_row_merges(worksheet, row: int) -> None:
    for start_col, end_col in TEST_MERGES:
        start = f"{get_column_letter(start_col)}{row}"
        end = f"{get_column_letter(end_col)}{row}"
        cell_range = f"{start}:{end}"
        if cell_range not in {str(rng) for rng in worksheet.merged_cells.ranges}:
            worksheet.merge_cells(cell_range)


def _write_tests(worksheet, display_rows: list[dict[str, Any]]) -> None:
    test_index = 0
    for row_index, item in enumerate(display_rows, start=1):
        row = TEST_START_ROW + row_index - 1
        if item["kind"] == "group":
            _write_product_group_header(worksheet, row, item["label"])
            continue

        test_index += 1
        test = item["test"]
        test_name = format_test_name(test.name, test.standard)
        worksheet[f"B{row}"] = test_index
        worksheet[f"D{row}"] = (
            f"Type of Test: {test_name}\n"
            f"Test Method/Requirement:\n{test.requirements.strip()}"
        )
        worksheet[f"AE{row}"] = SAC_CODE
        worksheet[f"AI{row}"] = coerce_number(test.qty)
        worksheet[f"AO{row}"] = coerce_number(test.batch)
        worksheet[f"AT{row}"] = coerce_number(test.total_cost)

        text = worksheet[f"D{row}"].value
        line_count = text.count("\n") + 1
        wrapped_lines = max(0, len(text) // 90)
        worksheet.row_dimensions[row].height = max(72, min(409, 16 * (line_count + wrapped_lines)))


def build_test_display_rows(tests: list[TestItem], product_lookup: dict[str, ProductItem]) -> list[dict[str, Any]]:
    if not any(test.product_keys for test in tests):
        return [{"kind": "test", "test": test} for test in tests]

    _validate_product_assignments(tests, product_lookup)
    groups: list[tuple[tuple[str, ...], list[TestItem]]] = []
    group_index_by_key: dict[tuple[str, ...], int] = {}
    for test in tests:
        keys = tuple(key for key in test.product_keys if key in product_lookup)
        if keys not in group_index_by_key:
            group_index_by_key[keys] = len(groups)
            groups.append((keys, []))
        groups[group_index_by_key[keys]][1].append(test)

    rows: list[dict[str, Any]] = []
    for index, (keys, grouped_tests) in enumerate(groups, start=1):
        products_text = ", ".join(product_quote_label(product_lookup[key]) for key in keys)
        rows.append({"kind": "group", "label": f"1.{index}. {products_text}", "product_keys": list(keys)})
        rows.extend({"kind": "test", "test": test} for test in grouped_tests)
    return rows


def _validate_product_assignments(tests: list[TestItem], product_lookup: dict[str, ProductItem]) -> None:
    if not product_lookup:
        raise ValueError("Add products before assigning tests to products.")
    product_names = [product.name.strip() for product in product_lookup.values()]
    duplicates = sorted({name for name in product_names if product_names.count(name) > 1})
    if duplicates:
        raise ValueError(
            "Product names must be unique when assigning tests to products. Duplicate: "
            + ", ".join(duplicates)
        )
    missing_tests = [test.name or test.standard or "Untitled test" for test in tests if not test.product_keys]
    if missing_tests:
        raise ValueError("Every test must have at least one product selected when product-specific tests are used.")
    valid_keys = set(product_lookup)
    unknown = sorted({key for test in tests for key in test.product_keys if key not in valid_keys})
    if unknown:
        raise ValueError(f"These selected products are no longer in the product list: {', '.join(unknown)}")
    covered = {key for test in tests for key in test.product_keys}
    uncovered = [product.name for key, product in product_lookup.items() if key not in covered]
    if uncovered:
        raise ValueError(
            "Every product must have at least one assigned test. Missing: "
            + ", ".join(uncovered)
        )


def _write_product_group_header(worksheet, row: int, label: str) -> None:
    for merge_range in list(worksheet.merged_cells.ranges):
        cell_range = CellRange(str(merge_range))
        if cell_range.min_row == row and cell_range.max_row == row:
            worksheet.unmerge_cells(str(cell_range))
    for col in range(2, 51):
        worksheet.cell(row, col).value = ""
    worksheet.merge_cells(f"B{row}:AX{row}")
    worksheet[f"B{row}"] = label
    worksheet[f"B{row}"].font = Font(bold=True, size=11)
    worksheet[f"B{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    worksheet[f"B{row}"].fill = PatternFill(fill_type="solid", fgColor="FFFFFF")
    worksheet.row_dimensions[row].height = max(42, min(120, 18 * ((len(label) // 95) + 1)))


def _write_totals(
    worksheet,
    display_rows: list[dict[str, Any]],
    *,
    use_discount: bool,
    discount_percent: float,
    instate: bool,
) -> None:
    tests = [item["test"] for item in display_rows if item["kind"] == "test"]
    last_test_row = TEST_START_ROW + len(display_rows) - 1
    amount_row = last_test_row + 1
    total_row = amount_row
    subtotal_cell = f"AN{total_row}"

    if use_discount:
        discount_row = total_row + 1
        net_total_row = total_row + 2
        cgst_row = total_row + 3
        sgst_row = total_row + 4
        igst_row = total_row + 5
        grand_total_row = total_row + 6
        tax_base_cell = f"AN{net_total_row}"
    else:
        cgst_row = total_row + 1
        sgst_row = total_row + 2
        igst_row = total_row + 3
        grand_total_row = total_row + 4
        tax_base_cell = subtotal_cell

    grand_total = _estimated_grand_total(tests, use_discount, discount_percent, instate)

    worksheet[f"B{amount_row}"] = "Amount in Words"
    worksheet[f"H{amount_row}"] = amount_in_words(grand_total)
    worksheet[f"AE{total_row}"] = "Total"
    worksheet[f"AN{total_row}"] = f"=SUM(AT{TEST_START_ROW}:AX{last_test_row})"

    if use_discount:
        percent_text = _format_percent(discount_percent)
        worksheet[f"AE{discount_row}"] = f"Discount - {percent_text}%"
        worksheet[f"AN{discount_row}"] = f"={subtotal_cell}*{discount_percent}%"
        worksheet[f"AE{net_total_row}"] = "New Total"
        worksheet[f"AN{net_total_row}"] = f"={subtotal_cell}-AN{discount_row}"

    worksheet[f"AE{cgst_row}"] = "CGST 9%" if instate else "CGST 0%"
    worksheet[f"AN{cgst_row}"] = f"=9%*{tax_base_cell}" if instate else 0
    worksheet[f"AE{sgst_row}"] = "SGST 9%" if instate else "SGST 0%"
    worksheet[f"AN{sgst_row}"] = f"=9%*{tax_base_cell}" if instate else 0
    worksheet[f"AE{igst_row}"] = "IGST 0%" if instate else "IGST 18%"
    worksheet[f"AN{igst_row}"] = 0 if instate else f"=18%*{tax_base_cell}"
    worksheet[f"AE{grand_total_row}"] = "Grand Total"
    worksheet[f"AN{grand_total_row}"] = (
        f"={tax_base_cell}+AN{cgst_row}+AN{sgst_row}+AN{igst_row}"
    )
    _ensure_below_test_rows_visible(worksheet, grand_total_row + 1)


def _format_percent(value: float) -> str:
    if float(value).is_integer():
        return str(int(value))
    return f"{value:g}"


def _estimated_grand_total(
    tests: list[TestItem],
    use_discount: bool = False,
    discount_percent: float = 0.0,
    instate: bool = True,
) -> Decimal | None:
    total = Decimal("0")
    found_any = False
    for test in tests:
        value = str(test.total_cost).strip().replace(",", "")
        if not value or value.startswith("="):
            continue
        try:
            total += Decimal(value)
            found_any = True
        except InvalidOperation:
            pass
    if not found_any:
        return None
    discount_amount = total * (Decimal(str(discount_percent)) / Decimal("100")) if use_discount else Decimal("0")
    net_total = total - discount_amount
    tax = net_total * Decimal("0.18")
    return (net_total + tax).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def _shift_images_for_insert(worksheet, insert_at: int, amount: int) -> None:
    for image in getattr(worksheet, "_images", []):
        anchor = getattr(image, "anchor", None)
        marker = getattr(anchor, "_from", None)
        if marker is not None and marker.row + 1 >= insert_at:
            marker.row += amount
        to_marker = getattr(anchor, "to", None)
        if to_marker is not None and to_marker.row + 1 >= insert_at:
            to_marker.row += amount


def _ensure_below_test_rows_visible(worksheet, start_row: int) -> None:
    for row in range(start_row, min(worksheet.max_row, start_row + 80) + 1):
        row_text = " ".join(
            str(worksheet.cell(row, col).value)
            for col in range(1, min(worksheet.max_column, 50) + 1)
            if worksheet.cell(row, col).value not in (None, "")
        )
        if not row_text:
            continue
        for cell in worksheet[row]:
            if cell.value not in (None, "") and cell.alignment:
                cell.alignment = copy(cell.alignment.copy(wrap_text=True, vertical="top"))
        current_height = worksheet.row_dimensions[row].height or 14.25
        estimated_lines = max(row_text.count("\n") + 1, (len(row_text) // 115) + 1)
        if len(row_text) > 40:
            worksheet.row_dimensions[row].height = max(current_height, min(90, 14.25 * estimated_lines))


def _write_payment_terms(worksheet, payment_terms: str) -> None:
    row = _find_payment_terms_row(worksheet)
    if row is None:
        return
    terms = payment_terms.strip() or "As per agreement with the Customer"
    worksheet[f"H{row}"] = terms


def _find_payment_terms_row(worksheet) -> int | None:
    for row in range(1, worksheet.max_row + 1):
        if _cell_text(worksheet[f"B{row}"].value).lower() == "payment terms":
            return row
    return None


def _ensure_signature_stamp(worksheet) -> None:
    payment_row = _find_payment_terms_row(worksheet)
    if payment_row is None or not SIGNATURE_STAMP_PATH.exists():
        return

    target_row = max(payment_row - 3, 1)
    target_col = 42
    kept_images = []
    for image in getattr(worksheet, "_images", []):
        marker = getattr(getattr(image, "anchor", None), "_from", None)
        if marker is not None and marker.row + 1 >= target_row - 2 and marker.col + 1 >= 35:
            continue
        kept_images.append(image)
    worksheet._images = kept_images

    stamp = Image(str(SIGNATURE_STAMP_PATH))
    stamp.width = 62
    stamp.height = 62
    stamp.anchor = f"{get_column_letter(target_col)}{target_row}"
    worksheet.add_image(stamp)


def load_quote(path: str | Path) -> tuple[dict[str, str], list[TestItem], list[ProductItem]]:
    workbook = load_workbook(path, data_only=False)
    worksheet = workbook.active
    quote_ref = _cell_text(worksheet["AJ3"].value)
    if not re.fullmatch(rf"{QUOTE_PREFIX}[A-Za-z0-9_-]+", quote_ref):
        raise ValueError("This does not look like a generated SKC quote workbook.")
    data = {
        "quote_suffix": normalize_quote_suffix(quote_ref) if quote_ref.upper().startswith(QUOTE_PREFIX) else quote_ref,
        "quote_date": _cell_text(worksheet["AR4"].value),
        "email_date": _cell_text(worksheet["X10"].value),
        "quote_made_by": "",
        "technical_engineers": "",
        "discount_enabled": "",
        "discount_percent": "",
        "instate": "",
        "payment_terms": "",
    }
    data.update(_parse_customer_details(_cell_text(worksheet["AJ6"].value)))
    data.update(_parse_agency_details(_cell_text(worksheet["B6"].value)))
    products = _parse_product_details(_cell_text(worksheet["L6"].value))

    total_row = _find_label_row(worksheet, "Total", start_row=TEST_START_ROW)
    if total_row is None:
        raise ValueError("Could not find the Total row in this quote.")
    tests = _parse_tests(worksheet, total_row, products)

    discount_row = total_row + 1
    discount_label = _cell_text(worksheet[f"AE{discount_row}"].value)
    if discount_label.lower().startswith("discount"):
        data["discount_enabled"] = "1"
        match = re.search(r"([\d.]+)", discount_label)
        if match:
            data["discount_percent"] = match.group(1)

    tax_labels = [
        _cell_text(worksheet[f"AE{row}"].value).upper()
        for row in range(total_row + 1, min(total_row + 8, worksheet.max_row + 1))
    ]
    data["instate"] = "1" if any("CGST 9" in label for label in tax_labels) and any("SGST 9" in label for label in tax_labels) else ""
    payment_row = _find_payment_terms_row(worksheet)
    if payment_row is not None:
        data["payment_terms"] = _cell_text(worksheet[f"H{payment_row}"].value)
    return data, tests, products


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _find_label_row(worksheet, label: str, start_row: int = 1) -> int | None:
    target = label.strip().lower()
    for row in range(start_row, worksheet.max_row + 1):
        if _cell_text(worksheet[f"AE{row}"].value).lower() == target:
            return row
    return None


def _parse_tests(worksheet, total_row: int, products: list[ProductItem] | None = None) -> list[TestItem]:
    tests = []
    current_product_keys: list[str] = []
    products = products or []
    for row in range(TEST_START_ROW, total_row):
        text = _cell_text(worksheet[f"D{row}"].value)
        if not text:
            group_keys = _parse_product_group_keys(_cell_text(worksheet[f"B{row}"].value), products)
            if group_keys:
                current_product_keys = group_keys
            continue
        name = text
        standard = ""
        requirements = ""
        match = re.search(
            r"Type of Test:\s*(.*?)\s*(?:\n|\r\n)Test Method/Requirement:\s*(.*)",
            text,
            flags=re.DOTALL,
        )
        if match:
            name, standard = split_test_name_standard(match.group(1).strip())
            requirements = match.group(2).strip()
        tests.append(
            TestItem(
                name=name,
                requirements=requirements,
                qty=_cell_text(worksheet[f"AI{row}"].value) or "1",
                batch=_cell_text(worksheet[f"AO{row}"].value) or "1",
                total_cost=_cell_text(worksheet[f"AT{row}"].value),
                standard=standard,
                product_keys=list(current_product_keys),
            )
        )
    return tests


def _parse_product_group_keys(label: str, products: list[ProductItem]) -> list[str]:
    if not re.match(r"^1\.\d+\.", label.strip()):
        return []
    return [product_key(product) for product in products if product_quote_label(product) in label]


def format_test_name(name: str, standard: str = "") -> str:
    name = name.strip()
    standard = standard.strip()
    if name and standard:
        return f"{name} - {standard}"
    return name or standard


def split_test_name_standard(value: str) -> tuple[str, str]:
    text = value.strip()
    if " - " not in text:
        return text, ""
    name, standard = text.rsplit(" - ", 1)
    return name.strip(), standard.strip()


def _parse_customer_details(text: str) -> dict[str, str]:
    result = {"customer": "", "designation": "", "company": "", "tel_no": "", "email": "", "address": ""}
    labels = {
        "Customer": "customer",
        "Designation": "designation",
        "Company": "company",
        "Tel No.": "tel_no",
        "E-mail": "email",
        "Address": "address",
    }
    for line in text.splitlines():
        for label, key in labels.items():
            prefix = f"{label}:"
            if line.startswith(prefix):
                result[key] = line[len(prefix) :].strip()
                break
    return result


def _parse_agency_details(text: str) -> dict[str, str]:
    result = {"quote_made_by": "", "technical_engineers": ""}
    for line in text.splitlines():
        if line.strip().startswith("Quote Made By:"):
            result["quote_made_by"] = line.split(":", 1)[1].strip()
        elif line.strip().startswith("Technical Engineer(s):"):
            result["technical_engineers"] = line.split(":", 1)[1].strip()
    return result


def _parse_product_details(text: str) -> list[ProductItem]:
    products = []
    for block in re.split(r"\n\s*\n", text.strip()):
        if not block.strip():
            continue
        fields = {"name": "", "dimensions": "", "weight": ""}
        for line in block.splitlines():
            if line.startswith("Product Name:"):
                fields["name"] = line.split(":", 1)[1].strip()
            elif line.startswith("Dimensions:"):
                fields["dimensions"] = line.split(":", 1)[1].strip()
            elif line.startswith("Weight:"):
                fields["weight"] = line.split(":", 1)[1].strip()
        if fields["name"]:
            dim1, dim2, dim3, dim_unit = _parse_dimensions(fields["dimensions"])
            weight, weight_unit = _parse_value_with_unit(fields["weight"], "kg")
            products.append(ProductItem(fields["name"], dim1, dim2, dim3, dim_unit, weight, weight_unit))
    return products


def _parse_dimensions(text: str) -> tuple[str, str, str, str]:
    value, unit = _parse_value_with_unit(text, "mm")
    parts = [part.strip() for part in value.split("x")]
    while len(parts) < 3:
        parts.append("")
    return parts[0], parts[1], parts[2], unit


def _parse_value_with_unit(text: str, default_unit: str) -> tuple[str, str]:
    text = text.strip()
    if not text:
        return "", default_unit
    pieces = text.rsplit(" ", 1)
    if len(pieces) == 2 and pieces[1].isalpha():
        return pieces[0].strip(), pieces[1].strip()
    return text, default_unit


ONES = [
    "",
    "one",
    "two",
    "three",
    "four",
    "five",
    "six",
    "seven",
    "eight",
    "nine",
    "ten",
    "eleven",
    "twelve",
    "thirteen",
    "fourteen",
    "fifteen",
    "sixteen",
    "seventeen",
    "eighteen",
    "nineteen",
]
TENS = ["", "", "twenty", "thirty", "forty", "fifty", "sixty", "seventy", "eighty", "ninety"]


def amount_in_words(amount: float | Decimal | None) -> str:
    if amount is None:
        return ""
    decimal_amount = Decimal(str(amount)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    rupees = int(decimal_amount)
    paise = int(((decimal_amount - Decimal(rupees)) * 100).quantize(Decimal("1"), rounding=ROUND_HALF_UP))
    if paise == 100:
        rupees += 1
        paise = 0
    if rupees <= 0 and paise <= 0:
        return "zero rupees only"
    parts = []
    if rupees:
        rupee_label = "rupee" if rupees == 1 else "rupees"
        parts.append(f"{_indian_number_words(rupees)} {rupee_label}")
    if paise:
        parts.append(f"{_indian_number_words(paise)} paise")
    return f"{' and '.join(parts)} only".capitalize()


def _under_thousand(value: int) -> str:
    parts = []
    hundreds, remainder = divmod(value, 100)
    if hundreds:
        parts.append(f"{ONES[hundreds]} hundred")
    if remainder:
        if remainder < 20:
            parts.append(ONES[remainder])
        else:
            tens, ones = divmod(remainder, 10)
            parts.append(TENS[tens] + (f" {ONES[ones]}" if ones else ""))
    return " ".join(parts)


def _indian_number_words(value: int) -> str:
    crore, value = divmod(value, 10_000_000)
    lakh, value = divmod(value, 100_000)
    thousand, value = divmod(value, 1_000)
    parts = []
    if crore:
        parts.append(f"{_under_thousand(crore)} crore")
    if lakh:
        parts.append(f"{_under_thousand(lakh)} lakh")
    if thousand:
        parts.append(f"{_under_thousand(thousand)} thousand")
    if value:
        parts.append(_under_thousand(value))
    return re.sub(r"\s+", " ", " ".join(parts)).strip()
